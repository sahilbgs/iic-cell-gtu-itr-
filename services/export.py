"""
GTU-ITR R&D & IIC Portal - Export Service
PDF export via ReportLab (SimpleDocTemplate + Platypus) and
Excel export via openpyxl. Both return in-memory BytesIO buffers.
"""

import io
import logging
from datetime import date, datetime

logger = logging.getLogger(__name__)


# ═══════════════════════════════════════════════════════════════════════════
# PDF Exporter
# ═══════════════════════════════════════════════════════════════════════════

class PDFExporter:
    """
    Generates PDF documents using ReportLab's Platypus framework.
    All methods return an ``io.BytesIO`` buffer ready for download,
    or ``None`` on failure.
    """

    # ------------------------------------------------------------------
    # Shared helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _get_styles():
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT

        styles = getSampleStyleSheet()

        styles.add(ParagraphStyle(
            name='DocTitle',
            parent=styles['Title'],
            fontSize=18,
            spaceAfter=20,
            alignment=TA_CENTER,
        ))
        styles.add(ParagraphStyle(
            name='SectionHeading',
            parent=styles['Heading2'],
            fontSize=13,
            spaceAfter=8,
            spaceBefore=14,
        ))
        styles.add(ParagraphStyle(
            name='MetaInfo',
            parent=styles['Normal'],
            fontSize=9,
            textColor='grey',
            alignment=TA_CENTER,
            spaceAfter=14,
        ))
        return styles

    @staticmethod
    def _build_table(data, col_widths=None):
        from reportlab.platypus import Table, TableStyle
        from reportlab.lib import colors

        table = Table(data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f6fa')]),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ]))
        return table

    @classmethod
    def _create_doc(cls, buffer, title="Report"):
        from reportlab.platypus import SimpleDocTemplate
        from reportlab.lib.pagesizes import A4

        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            title=title,
            topMargin=50,
            bottomMargin=50,
            leftMargin=40,
            rightMargin=40,
        )
        return doc

    # ------------------------------------------------------------------
    # Export: Proposals list
    # ------------------------------------------------------------------

    @classmethod
    def export_proposals(cls, proposals) -> io.BytesIO | None:
        """
        Export a list of Proposal model instances to a PDF table.
        """
        try:
            from reportlab.platypus import Paragraph, Spacer

            buffer = io.BytesIO()
            doc = cls._create_doc(buffer, "Proposals Report")
            styles = cls._get_styles()
            elements = []

            elements.append(Paragraph("Research Proposals Report", styles['DocTitle']))
            elements.append(Paragraph(
                f"Generated on {datetime.now().strftime('%d %B %Y, %I:%M %p')}",
                styles['MetaInfo'],
            ))

            # Table
            header = ['#', 'Title', 'Faculty', 'Department', 'Status', 'Budget (₹)']
            data = [header]
            for i, p in enumerate(proposals, 1):
                data.append([
                    str(i),
                    p.title[:60],
                    p.faculty.full_name if p.faculty else 'N/A',
                    p.department.name if p.department else 'N/A',
                    p.status_label,
                    f"{p.budget_amount:,.2f}" if p.budget_amount else 'N/A',
                ])

            col_widths = [25, 160, 100, 80, 65, 75]
            elements.append(cls._build_table(data, col_widths))
            elements.append(Spacer(1, 12))

            # Summary
            elements.append(Paragraph(
                f"Total Proposals: {len(proposals)}",
                styles['Normal'],
            ))

            doc.build(elements)
            buffer.seek(0)
            return buffer

        except Exception as exc:
            logger.exception("export_proposals PDF failed: %s", exc)
            return None

    # ------------------------------------------------------------------
    # Export: Schemes list
    # ------------------------------------------------------------------

    @classmethod
    def export_schemes(cls, schemes) -> io.BytesIO | None:
        """
        Export a list of Scheme model instances to a PDF table.
        """
        try:
            from reportlab.platypus import Paragraph, Spacer

            buffer = io.BytesIO()
            doc = cls._create_doc(buffer, "Schemes Report")
            styles = cls._get_styles()
            elements = []

            elements.append(Paragraph("Funding Schemes Report", styles['DocTitle']))
            elements.append(Paragraph(
                f"Generated on {datetime.now().strftime('%d %B %Y, %I:%M %p')}",
                styles['MetaInfo'],
            ))

            header = ['#', 'Title', 'Agency', 'Category', 'Deadline', 'Amount (₹)', 'Status']
            data = [header]
            for i, s in enumerate(schemes, 1):
                data.append([
                    str(i),
                    s.title[:50],
                    (s.funding_agency or 'N/A')[:30],
                    s.category_label,
                    s.deadline.strftime('%d-%m-%Y') if s.deadline else 'N/A',
                    f"{s.funding_amount:,.0f}" if s.funding_amount else 'N/A',
                    s.status_label,
                ])

            col_widths = [22, 130, 80, 65, 60, 70, 55]
            elements.append(cls._build_table(data, col_widths))
            elements.append(Spacer(1, 12))
            elements.append(Paragraph(
                f"Total Schemes: {len(schemes)}",
                styles['Normal'],
            ))

            doc.build(elements)
            buffer.seek(0)
            return buffer

        except Exception as exc:
            logger.exception("export_schemes PDF failed: %s", exc)
            return None

    # ------------------------------------------------------------------
    # Export: Report (narrative text)
    # ------------------------------------------------------------------

    @classmethod
    def export_report(cls, report) -> io.BytesIO | None:
        """
        Export a Report model instance (with narrative content) to PDF.
        """
        try:
            from reportlab.platypus import Paragraph, Spacer

            buffer = io.BytesIO()
            doc = cls._create_doc(buffer, report.title)
            styles = cls._get_styles()
            elements = []

            elements.append(Paragraph(report.title, styles['DocTitle']))
            elements.append(Paragraph(
                f"Period: {report.period_label} &nbsp;|&nbsp; "
                f"Type: {report.report_type_label} &nbsp;|&nbsp; "
                f"Generated: {report.created_at.strftime('%d %B %Y')}",
                styles['MetaInfo'],
            ))

            # Render content paragraphs
            content = report.content or report.summary or "No content available."
            for paragraph in content.split('\n'):
                paragraph = paragraph.strip()
                if not paragraph:
                    elements.append(Spacer(1, 6))
                    continue

                # Detect section headings (lines that are ALL-CAPS or start with a number)
                if paragraph.isupper() or (
                    len(paragraph) < 100 and paragraph[0].isdigit() and '.' in paragraph[:4]
                ):
                    elements.append(Paragraph(paragraph, styles['SectionHeading']))
                else:
                    elements.append(Paragraph(paragraph, styles['Normal']))

            doc.build(elements)
            buffer.seek(0)
            return buffer

        except Exception as exc:
            logger.exception("export_report PDF failed: %s", exc)
            return None

    # ------------------------------------------------------------------
    # Export: Publications
    # ------------------------------------------------------------------

    @classmethod
    def export_publications(cls, publications) -> io.BytesIO | None:
        """
        Export a list of Publication model instances to a PDF table.
        """
        try:
            from reportlab.platypus import Paragraph, Spacer

            buffer = io.BytesIO()
            doc = cls._create_doc(buffer, "Publications Report")
            styles = cls._get_styles()
            elements = []

            elements.append(Paragraph("Publications &amp; Patents Report", styles['DocTitle']))
            elements.append(Paragraph(
                f"Generated on {datetime.now().strftime('%d %B %Y, %I:%M %p')}",
                styles['MetaInfo'],
            ))

            header = ['#', 'Title', 'Authors', 'Type', 'Journal/Conf.', 'Impact Factor']
            data = [header]
            for i, pub in enumerate(publications, 1):
                data.append([
                    str(i),
                    pub.title[:50],
                    pub.authors[:40],
                    pub.pub_type_label,
                    (pub.journal or 'N/A')[:30],
                    str(pub.impact_factor) if pub.impact_factor else 'N/A',
                ])

            col_widths = [22, 140, 100, 55, 100, 55]
            elements.append(cls._build_table(data, col_widths))
            elements.append(Spacer(1, 12))
            elements.append(Paragraph(
                f"Total Publications: {len(publications)}",
                styles['Normal'],
            ))

            doc.build(elements)
            buffer.seek(0)
            return buffer

        except Exception as exc:
            logger.exception("export_publications PDF failed: %s", exc)
            return None


# ═══════════════════════════════════════════════════════════════════════════
# Excel Exporter
# ═══════════════════════════════════════════════════════════════════════════

class ExcelExporter:
    """
    Generates Excel workbooks using openpyxl.
    All methods return an ``io.BytesIO`` buffer, or ``None`` on failure.
    """

    # ------------------------------------------------------------------
    # Shared helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _style_header(ws, row=1):
        """Apply header styling to the first row."""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        for cell in ws[row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border

    @staticmethod
    def _auto_width(ws):
        """Auto-size column widths (approximation)."""
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    @staticmethod
    def _save_to_buffer(wb) -> io.BytesIO:
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    # ------------------------------------------------------------------
    # Export: Proposals
    # ------------------------------------------------------------------

    @classmethod
    def export_proposals(cls, proposals) -> io.BytesIO | None:
        """Export proposals to an Excel workbook."""
        try:
            from openpyxl import Workbook

            wb = Workbook()
            ws = wb.active
            ws.title = "Proposals"

            headers = [
                'Sr.', 'Title', 'Faculty', 'Department', 'Scheme',
                'Status', 'Budget (₹)', 'Submitted On', 'Created On',
            ]
            ws.append(headers)
            cls._style_header(ws)

            for i, p in enumerate(proposals, 1):
                ws.append([
                    i,
                    p.title,
                    p.faculty.full_name if p.faculty else 'N/A',
                    p.department.name if p.department else 'N/A',
                    p.scheme.title if p.scheme else 'N/A',
                    p.status_label,
                    p.budget_amount,
                    p.submitted_at.strftime('%d-%m-%Y') if p.submitted_at else '',
                    p.created_at.strftime('%d-%m-%Y') if p.created_at else '',
                ])

            cls._auto_width(ws)
            return cls._save_to_buffer(wb)

        except Exception as exc:
            logger.exception("export_proposals Excel failed: %s", exc)
            return None

    # ------------------------------------------------------------------
    # Export: Schemes
    # ------------------------------------------------------------------

    @classmethod
    def export_schemes(cls, schemes) -> io.BytesIO | None:
        """Export schemes to an Excel workbook."""
        try:
            from openpyxl import Workbook

            wb = Workbook()
            ws = wb.active
            ws.title = "Schemes"

            headers = [
                'Sr.', 'Title', 'Agency', 'Category', 'Deadline',
                'Amount (₹)', 'Status', 'Priority', 'Created On',
            ]
            ws.append(headers)
            cls._style_header(ws)

            for i, s in enumerate(schemes, 1):
                ws.append([
                    i,
                    s.title,
                    s.funding_agency or 'N/A',
                    s.category_label,
                    s.deadline.strftime('%d-%m-%Y') if s.deadline else '',
                    s.funding_amount,
                    s.status_label,
                    s.priority_label,
                    s.created_at.strftime('%d-%m-%Y') if s.created_at else '',
                ])

            cls._auto_width(ws)
            return cls._save_to_buffer(wb)

        except Exception as exc:
            logger.exception("export_schemes Excel failed: %s", exc)
            return None

    # ------------------------------------------------------------------
    # Export: Report data
    # ------------------------------------------------------------------

    @classmethod
    def export_report(cls, report) -> io.BytesIO | None:
        """Export a report's content to an Excel workbook."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font

            wb = Workbook()
            ws = wb.active
            ws.title = "Report"

            # Title row
            ws.append([report.title])
            ws.merge_cells('A1:D1')
            ws['A1'].font = Font(size=14, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')

            # Metadata
            ws.append(['Period', report.period_label])
            ws.append(['Type', report.report_type_label])
            ws.append(['Generated', report.created_at.strftime('%d %B %Y')])
            ws.append([])  # blank row

            # Content
            ws.append(['Report Content'])
            ws['A6'].font = Font(size=12, bold=True)

            content = report.content or report.summary or "No content."
            for line in content.split('\n'):
                ws.append([line])

            cls._auto_width(ws)
            return cls._save_to_buffer(wb)

        except Exception as exc:
            logger.exception("export_report Excel failed: %s", exc)
            return None

    # ------------------------------------------------------------------
    # Export: Publications
    # ------------------------------------------------------------------

    @classmethod
    def export_publications(cls, publications) -> io.BytesIO | None:
        """Export publications to an Excel workbook."""
        try:
            from openpyxl import Workbook

            wb = Workbook()
            ws = wb.active
            ws.title = "Publications"

            headers = [
                'Sr.', 'Title', 'Authors', 'Type', 'Journal/Conference',
                'DOI', 'ISSN', 'Impact Factor', 'Citations',
                'Published Date', 'Department',
            ]
            ws.append(headers)
            cls._style_header(ws)

            for i, pub in enumerate(publications, 1):
                ws.append([
                    i,
                    pub.title,
                    pub.authors,
                    pub.pub_type_label,
                    pub.journal or 'N/A',
                    pub.doi or '',
                    pub.issn or '',
                    pub.impact_factor,
                    pub.citation_count,
                    pub.published_date.strftime('%d-%m-%Y') if pub.published_date else '',
                    pub.department.name if pub.department else 'N/A',
                ])

            cls._auto_width(ws)
            return cls._save_to_buffer(wb)

        except Exception as exc:
            logger.exception("export_publications Excel failed: %s", exc)
            return None
